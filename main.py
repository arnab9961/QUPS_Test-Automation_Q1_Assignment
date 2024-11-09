from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
import re

def setup_driver():
    driver = webdriver.Edge()
    driver.get('https://www.google.com')
    return driver

def clean_html(text):
    # Remove HTML tags and clean up the text
    text = re.sub(r'<[^>]+>', '', text)
    text = text.strip()
    return text

def get_suggestions(driver, keyword):
    try:
        search = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, 'q'))
        )
        search.clear()
        
        for char in keyword:
            search.send_keys(char)
            time.sleep(0.2)
        
        time.sleep(1)
        
        suggestions = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(
                (By.XPATH, "//li[@role='presentation']//div[@class='wM6W7d']")
            )
        )
        
        texts = []
        for s in suggestions:
            text = clean_html(s.get_attribute('innerHTML'))
            if text and text.lower() != keyword.lower():
                texts.append(text)
        
        longest = max(texts, key=len) if texts else ''
        shortest = min(texts, key=len) if texts else ''
        
        return longest, shortest
    except Exception as e:
        print(f"Error getting suggestions for {keyword}: {str(e)}")
        return '', ''
    finally:
        driver.get('https://www.google.com')
        time.sleep(1)

def find_column_index(headers, keywords):
    """Find column index based on multiple possible header names"""
    for i, header in enumerate(headers):
        if header and any(keyword.lower() in str(header).lower() for keyword in keywords):
            return i
    return None

def process_sheet(sheet, driver):
    data = list(sheet.values)
    if not data:
        return
    
    # Find header row
    header_row = None
    for i, row in enumerate(data):
        if row and any(isinstance(cell, str) and 'keyword' in str(cell).lower() for cell in row):
            header_row = i
            break
    
    if header_row is None:
        print(f"No header row found in sheet {sheet.title}")
        return
    
    headers = list(data[header_row])
    
    # Find column indices
    keyword_col = find_column_index(headers, ['keyword'])
    longest_col = find_column_index(headers, ['longest', 'longest option'])
    shortest_col = find_column_index(headers, ['shortest', 'shortest option'])
    
    if keyword_col is None:
        print(f"No keyword column found in sheet {sheet.title}")
        return
    
    # Create columns if they don't exist
    if longest_col is None:
        headers.append('longest option')
        longest_col = len(headers) - 1
        sheet.cell(row=header_row + 1, column=longest_col + 1, value='longest option')
    
    if shortest_col is None:
        headers.append('shortest option')
        shortest_col = len(headers) - 1
        sheet.cell(row=header_row + 1, column=shortest_col + 1, value='shortest option')
    
    # Process keywords
    for row_idx, row in enumerate(data[header_row + 1:], header_row + 2):
        if row and len(row) > keyword_col:
            keyword = str(row[keyword_col]).strip()
            if keyword and keyword.lower() != 'none':
                print(f"Processing keyword: {keyword}")
                longest, shortest = get_suggestions(driver, keyword)
                
                # Write results to correct columns
                sheet.cell(row=row_idx, column=longest_col + 1, value=longest)
                sheet.cell(row=row_idx, column=shortest_col + 1, value=shortest)
                
                # Save after each keyword to prevent data loss
                try:
                    wb = sheet.parent
                    wb.save(wb.path)
                except Exception as e:
                    print(f"Error saving workbook: {str(e)}")

def main():
    file_path = r"C:\Users\dmraf\Downloads\Excel.xlsx"
    wb = load_workbook(file_path)
    driver = setup_driver()
    
    try:
        for sheet_name in wb.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            sheet = wb[sheet_name]
            process_sheet(sheet, driver)
            wb.save(file_path)
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
